"""
Student Score Management System

A comprehensive Flask web application for managing student academic records,
including user authentication, score tracking, grading, and reporting features.

Author: OSondu Stanley
Version: 1.0.0
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, IntegerField, TextAreaField, SelectField, validators
from flask_wtf.csrf import CSRFProtect
import sqlite3
import json
import csv
import re
from io import StringIO, BytesIO
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash

import os

import logging

# Set up logging
logging.basicConfig(filename='app.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

def init_db():
    """Initialize the SQLite database and create tables if they don't exist."""
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()

    # Create users table
    c.execute('''CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT DEFAULT 'teacher',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )''')

    # Add role column if it doesn't exist
    try:
        c.execute('ALTER TABLE users ADD COLUMN role TEXT DEFAULT "teacher"')
    except sqlite3.OperationalError:
        pass  # Column already exists

    # Create students table
    c.execute('''CREATE TABLE IF NOT EXISTS students (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id TEXT NOT NULL,
                    student_id TEXT NOT NULL,
                    firstname TEXT NOT NULL,
                    classname TEXT NOT NULL,
                    number_of_subject INTEGER NOT NULL,
                    subjects TEXT NOT NULL,
                    scores TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(user_id, student_id)
                )''')

    # Create reports table
    c.execute('''CREATE TABLE IF NOT EXISTS reports (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id TEXT NOT NULL,
                    description TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    status TEXT DEFAULT 'unread',
                    read_at TEXT
                )''')

    conn.commit()
    conn.close()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your_secret_key_here')  # Use environment variable for production

# Initialize database
init_db()

def load_users():
    """
    Load user credentials from SQLite database.

    Returns:
        dict: Dictionary of username -> {'password_hash': hash, 'role': role}
    """
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()
    try:
        c.execute('SELECT username, password_hash, role FROM users')
        users = {row[0]: {'password_hash': row[1], 'role': row[2] or 'teacher'} for row in c.fetchall()}
    except sqlite3.OperationalError:
        # Fallback if role column doesn't exist
        c.execute('SELECT username, password_hash FROM users')
        users = {row[0]: {'password_hash': row[1], 'role': 'teacher'} for row in c.fetchall()}
    conn.close()
    return users

def save_users(users):
    """
    Save user credentials to SQLite database.

    Args:
        users (dict): Dictionary of username -> {'password_hash': hash, 'role': role}
    """
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()
    c.execute('DELETE FROM users')  # Clear existing users
    for username, data in users.items():
        try:
            c.execute('INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                     (username, data['password_hash'], data.get('role', 'teacher')))
        except sqlite3.OperationalError:
            # Fallback if role column doesn't exist
            c.execute('INSERT INTO users (username, password_hash) VALUES (?, ?)',
                     (username, data['password_hash']))
    conn.commit()
    conn.close()

def hash_password(password):
    """
    Hash a password using Werkzeug security.

    Args:
        password (str): Plain text password

    Returns:
        str: Hashed password
    """
    return generate_password_hash(password)

def check_password(hashed, password):
    """
    Verify a password against its hash.

    Args:
        hashed (str): Hashed password
        password (str): Plain text password to check

    Returns:
        bool: True if password matches hash
    """
    return check_password_hash(hashed, password)

def get_valid_score(prompt, minimum, maximum):
    """Prompts the user for a valid score within a specified range."""
    while True:
        try:
            score = int(input(prompt))
            if minimum <= score <= maximum:
                return score
            else:
                print(f"Invalid input! Enter a value between {minimum} and {maximum}.")
        except ValueError:
            print("Invalid input! Please enter a valid integer.")

def rank_students_by_average(students):
    valid_students = [
        s for s in students
        if isinstance(s, dict) and "average_marks" in s
    ]
    return sorted(valid_students, key=lambda x: x["average_marks"], reverse=True)

def calculate_positions(students_list):
    positions = {}
    class_groups = {}
    for student in students_list:
        class_name = student['class_name']
        if class_name not in class_groups:
            class_groups[class_name] = []
        class_groups[class_name].append(student)

    for class_name, class_students in class_groups.items():
        sorted_students = sorted(class_students, key=lambda x: x['average_marks'], reverse=True)
        for i, student in enumerate(sorted_students, 1):
            positions[student['student_id']] = {
                'pos': i,
                'size': len(sorted_students),
                'class': class_name
            }
    return positions

def find_student_id(user_students, student_id):
    sid = student_id.strip().lower()
    for student in user_students:
        if not isinstance(student, dict):
            continue
        if student.get("student_id", "").strip().lower() == sid:
            return student
    return None

def id_exists(user_students, student_id):
    sid = student_id.strip().lower()
    return any(isinstance(student, dict) and student.get("student_id", "").strip().lower() == sid for student in user_students)

def load_students(user_id):
    """
    Load students for a specific user from SQLite database.

    Args:
        user_id (str): The user ID to load students for

    Returns:
        dict: Dictionary of student_id -> student_data
    """
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()
    c.execute('SELECT student_id, firstname, classname, number_of_subject, subjects, scores FROM students WHERE user_id = ?',
             (user_id,))
    students_data = {}
    for row in c.fetchall():
        student_id, firstname, classname, number_of_subject, subjects_str, scores_str = row
        subjects = json.loads(subjects_str) if subjects_str else []
        scores = json.loads(scores_str) if scores_str else {}
        students_data[student_id] = {
            'firstname': firstname,
            'classname': classname,
            'number_of_subject': number_of_subject,
            'subjects': subjects,
            'scores': scores
        }
    conn.close()
    return students_data

def save_student(user_id, student_id, student_data):
    """
    Save a student to SQLite database.

    Args:
        user_id (str): The user ID
        student_id (str): The student ID
        student_data (dict): The student data
    """
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()
    subjects_str = json.dumps(student_data['subjects'])
    scores_str = json.dumps(student_data['scores'])
    c.execute('''INSERT OR REPLACE INTO students
                 (user_id, student_id, firstname, classname, number_of_subject, subjects, scores)
                 VALUES (?, ?, ?, ?, ?, ?, ?)''',
             (user_id, student_id, student_data['firstname'], student_data['classname'],
              student_data['number_of_subject'], subjects_str, scores_str))
    conn.commit()
    conn.close()

def delete_student(user_id, student_id):
    """
    Delete a student from SQLite database.

    Args:
        user_id (str): The user ID
        student_id (str): The student ID
    """
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()
    c.execute('DELETE FROM students WHERE user_id = ? AND student_id = ?',
             (user_id, student_id))
    conn.commit()
    conn.close()

def load_reports():
    """
    Load reports from SQLite database.

    Returns:
        list: List of report dictionaries
    """
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()
    c.execute('SELECT id, user_id, description, timestamp, status, read_at FROM reports ORDER BY timestamp DESC')
    reports = []
    for row in c.fetchall():
        report_id, user_id, description, timestamp, status, read_at = row
        reports.append({
            'id': report_id,
            'user_id': user_id,
            'description': description,
            'timestamp': timestamp,
            'status': status,
            'read_at': read_at
        })
    conn.close()
    return reports

def save_report(report):
    """
    Save a report to SQLite database.

    Args:
        report (dict): Report dictionary
    """
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()
    c.execute('INSERT INTO reports (user_id, description, timestamp, status) VALUES (?, ?, ?, ?)',
             (report['user_id'], report['description'], report['timestamp'], report['status']))
    conn.commit()
    conn.close()

def mark_report_read(report_id):
    """
    Mark a report as read by updating the read_at timestamp.

    Args:
        report_id (int): The ID of the report to mark as read
    """
    conn = sqlite3.connect('student_score.db')
    c = conn.cursor()
    c.execute('UPDATE reports SET status = ?, read_at = ? WHERE id = ?',
             ('read', datetime.now().isoformat(), report_id))
    conn.commit()
    conn.close()

@app.route('/', methods=['GET'])
def home():
    return render_template('home.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        users = load_users()
        if username in users and check_password(users[username]['password_hash'], password):
            session['user_id'] = username
            session['role'] = users[username]['role']
            return redirect(url_for('menu'))
        else:
            flash('Invalid username or password')
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['email'].strip()
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        if password != confirm_password:
            flash('Passwords do not match')
            return redirect(url_for('signup'))
        users = load_users()
        if username in users:
            flash('Email already exists')
            return redirect(url_for('signup'))
        users[username] = {'password_hash': hash_password(password), 'role': 'teacher'}
        save_users(users)
        flash('Account created successfully. Please log in.')
        return redirect(url_for('login'))
    return render_template('signup.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('role', None)
    return redirect(url_for('home'))

@app.route('/menu')
def menu():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    role = session.get('role', 'teacher')
    if role == 'student':
        return redirect(url_for('student_menu'))
    return render_template('menu.html')

@app.route('/student_menu')
def student_menu():
    if 'user_id' not in session or session.get('role') != 'student':
        return redirect(url_for('login'))
    return render_template('student_menu.html')

@app.route('/my_report')
def my_report():
    if 'user_id' not in session or session.get('role') != 'student':
        return redirect(url_for('login'))

    student_id = session['user_id']
    user_id = 'admin'  # Assume students are under admin or something, but for now, assume user_id is the teacher, but since student logs in with student_id, perhaps user_id is the student_id.

    # For simplicity, assume students are stored under their own user_id = student_id.

    user_students = load_students(student_id)  # Wait, no.

    # To make it work, perhaps students have their own user_id = student_id, and they have one student with that ID.

    # So, user_students = load_students(student_id), and the student is student_id.

    student_data = load_students(student_id).get(student_id)

    if not student_data:
        flash('Student data not found.')
        return redirect(url_for('student_menu'))

    scores = student_data.get('scores', {})

    if scores:
        overall_marks = [subj_data.get('overall_mark', 0) for subj_data in scores.values() if isinstance(subj_data, dict)]
        average_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
        total_marks = sum(overall_marks)
    else:
        average_marks = 0
        total_marks = 0

    if average_marks >= 70:
        grade = 'A'
    elif average_marks >= 60:
        grade = 'B'
    elif average_marks >= 50:
        grade = 'C'
    elif average_marks >= 40:
        grade = 'D'
    else:
        grade = 'F'

    status = 'Pass' if average_marks >= 60 else 'Fail'

    subjects_with_scores = {}
    for subject in student_data.get('subjects', []):
        subj_scores = scores.get(subject, {})
        if isinstance(subj_scores, dict):
            subjects_with_scores[subject] = {
                'tests': subj_scores.get('total_test', 0),
                'exam': subj_scores.get('total_exam', 0),
                'total': subj_scores.get('overall_mark', 0)
            }
        else:
            subjects_with_scores[subject] = {'tests': 0, 'exam': 0, 'total': 0}

    # Positions: since single student, perhaps not needed, or calculate across all.

    student = {
        'first_name': student_data.get('firstname', ''),
        'student_id': student_id,
        'class_name': student_data.get('classname', ''),
        'number_of_subject': student_data.get('number_of_subject', 0),
        'subjects': subjects_with_scores,
        'total_marks': total_marks,
        'average_marks': average_marks,
        'Grade': grade,
        'Status': status
    }

    return render_template('student_report.html', student=student, position=None, 
                           chart_labels=list(student['subjects'].keys()), 
                           chart_data=[subj_data.get('overall_mark', 0) for subj_data in student['subjects'].values()])

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        users = load_users()
        user_id = session['user_id']

        if user_id not in users or not check_password(users[user_id]['password_hash'], current_password):
            flash('Current password is incorrect.', 'error')
            return redirect(url_for('change_password'))

        if new_password != confirm_password:
            flash('New passwords do not match.', 'error')
            return redirect(url_for('change_password'))

        if len(new_password) < 6:
            flash('New password must be at least 6 characters long.', 'error')
            return redirect(url_for('change_password'))

        users[user_id] = {'password_hash': hash_password(new_password), 'role': users[user_id].get('role', 'teacher')}
        save_users(users)
        flash('Password changed successfully!', 'success')
        return redirect(url_for('menu'))

    return render_template('change_password.html')

@app.route('/bulk_import', methods=['GET', 'POST'])
def bulk_import():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith('.csv'):
            flash('Please upload a valid CSV file.', 'error')
            return redirect(url_for('bulk_import'))

        user_id = session['user_id']
        user_students = load_students(user_id)

        csv_content = file.read().decode('utf-8')
        reader = csv.DictReader(StringIO(csv_content))

        imported = 0
        errors = []

        for row in reader:
            student_id = row.get('Student ID', '').strip()
            firstname = row.get('First Name', '').strip()
            classname = row.get('Class', '').strip()
            subjects_str = row.get('Subjects', '').strip()

            if not student_id or not firstname or not classname or not subjects_str:
                errors.append(f"Missing data for student {student_id}")
                continue

            subjects = [s.strip() for s in subjects_str.split(',') if s.strip()]
            number_of_subject = len(subjects)

            if student_id in user_students:
                errors.append(f"Student ID {student_id} already exists")
                continue

            student_data = {
                'firstname': firstname,
                'classname': classname,
                'number_of_subject': number_of_subject,
                'subjects': subjects,
                'scores': {}
            }

            save_student(user_id, student_id, student_data)
            imported += 1

        if imported > 0:
            flash(f'Successfully imported {imported} students.', 'success')
        if errors:
            for error in errors:
                flash(error, 'error')

        return redirect(url_for('view_students'))

    return render_template('bulk_import.html')

@app.route('/add_student', methods=['GET', 'POST'])
def add_student():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        student_id = request.form['student_id']
        firstname = request.form['firstname']
        classname = request.form['classname']
        number_of_subject = int(request.form['number_of_subject'])

        # Validation: Check student_id format (e.g., alphanumeric, no spaces)
        if not re.match(r'^[A-Za-z0-9_-]+$', student_id):
            flash("Student ID must contain only letters, numbers, underscores, or hyphens.", "error")
            return redirect(url_for('add_student'))

        # Validation: Check firstname format
        if not firstname or not re.match(r'^[A-Za-z\s]+$', firstname):
            flash("First name must contain only letters and spaces.", "error")
            return redirect(url_for('add_student'))

        # Validation: Check classname
        if not classname:
            flash("Class name is required.", "error")
            return redirect(url_for('add_student'))

        # Validation: Check if number of subjects matches
        if len(subjects) != number_of_subject:
            flash("Number of subjects does not match the subjects provided.", "error")
            return redirect(url_for('add_student'))

        # Validation: Check for duplicate subjects
        if len(set(subjects)) != len(subjects):
            flash("Duplicate subjects are not allowed.", "error")
            return redirect(url_for('add_student'))

        user_id = session['user_id']
        student_data = {
            'firstname': firstname,
            'classname': classname,
            'number_of_subject': number_of_subject,
            'subjects': subjects,
            'scores': {}
        }

        save_student(user_id, student_id, student_data)
        flash('Student added successfully! Now enter scores for each subject.', 'success')
        logging.info(f"User {user_id} added student {student_id}")
        return redirect(url_for('enter_scores', student_id=student_id))

    return render_template('add_student.html')

@app.route('/view_students')
def view_students():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    user_students = load_students(user_id)

    page = int(request.args.get('page', 1))
    per_page = 10  # Students per page

    # Transform data to match template expectations
    students_list = []
    total_average = 0
    grade_a_count = 0
    pass_count = 0

    for student_id, student_data in user_students.items():
        # Calculate average marks from overall_mark
        scores = student_data.get('scores', {})
        if scores:
            overall_marks = []
            for subj_data in scores.values():
                if isinstance(subj_data, dict):
                    # New format: detailed score object
                    overall_marks.append(subj_data.get('overall_mark', 0))
                elif isinstance(subj_data, (int, float)):
                    # Old format: simple number
                    overall_marks.append(float(subj_data))
            average_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
        else:
            average_marks = 0

        # Determine grade
        if average_marks >= 70:
            grade = 'A'
            grade_a_count += 1
        elif average_marks >= 60:
            grade = 'B'
        elif average_marks >= 50:
            grade = 'C'
        elif average_marks >= 40:
            grade = 'D'
        else:
            grade = 'F'

        # Determine status
        status = 'Pass' if average_marks >= 60 else 'Fail'
        if status == 'Pass':
            pass_count += 1

        total_average += average_marks

        students_list.append({
            'first_name': student_data.get('firstname', ''),
            'student_id': student_id,
            'class_name': student_data.get('classname', ''),
            'subjects': {
                subj: {
                    'overall_mark': subj_data.get('overall_mark', 0) if isinstance(subj_data, dict) else 0,
                    'grade': subj_data.get('grade', grade) if isinstance(subj_data, dict) else grade,
                    'status': status,
                    'tests': [subj_data.get('first_test', 0), subj_data.get('second_test', 0), subj_data.get('third_test', 0)] if isinstance(subj_data, dict) else [],
                    'cbt': subj_data.get('objective', 0) if isinstance(subj_data, dict) else 0,
                    'theory_exam': subj_data.get('theory', 0) if isinstance(subj_data, dict) else 0,
                    'total_exam': subj_data.get('total_exam', 0) if isinstance(subj_data, dict) else 0,
                    'total_test': subj_data.get('total_test', 0) if isinstance(subj_data, dict) else 0
                }
                for subj, subj_data in scores.items()
            },
            'average_marks': average_marks,
            'Grade': grade,
            'Status': status
        })

    # Calculate overall statistics
    overall_average = total_average / len(students_list) if students_list else 0

    # Get unique classes for filter dropdown
    classes = list(set(student['class_name'] for student in students_list))
    positions = calculate_positions(students_list)

    # Pagination
    total = len(students_list)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_students = students_list[start:end]

    return render_template('view_students.html',
                         students=paginated_students,
                         total_students=total,
                         grade_a_count=grade_a_count,
                         pass_count=pass_count,
                         overall_average=overall_average,
                         classes=classes,
                         positions=positions,
                         page=page,
                         per_page=per_page,
                         total_pages=(total + per_page - 1) // per_page)

@app.route('/edit_student', methods=['GET', 'POST'])
def edit_student():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    user_students = load_students(user_id)

    if request.method == 'POST':
        student_id = request.form.get('student_id')
        if student_id and student_id in user_students:
            # Update student data
            student_data = user_students[student_id]
            student_data['firstname'] = request.form.get('firstname', '')
            student_data['classname'] = request.form.get('classname', '')
            student_data['number_of_subject'] = int(request.form.get('number_of_subject', 0))
            student_data['subjects'] = [s.strip() for s in re.split(r'[,\s]+', request.form.get('subjects', '')) if s.strip()]

            save_student(user_id, student_id, student_data)
            flash('Student updated successfully!')
            return redirect(url_for('view_students'))
        else:
            flash('Student not found.')

    # For GET request, show form to select student
    student_id = request.args.get('student_id')
    if student_id and student_id in user_students:
        student = user_students[student_id]
        return render_template('edit_student.html', student=student, student_id=student_id, students=user_students)
    else:
        return render_template('edit_student.html', students=user_students)

@app.route('/delete_student', methods=['GET', 'POST'])
def delete_student():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    user_students = load_students(user_id)

    if request.method == 'POST':
        student_id = request.form.get('student_id')
        if student_id and student_id in user_students:
            delete_student(user_id, student_id)
            flash('Student deleted successfully!')
            return redirect(url_for('view_students'))
        else:
            flash('Student not found.')

    return render_template('delete_student.html', students=user_students)

@app.route('/search_student', methods=['GET', 'POST'])
def search_student():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    user_students = load_students(user_id)

    student = None
    positions = {}

    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip().lower()
        if student_id:
            for sid, sdata in user_students.items():
                if sid.lower() == student_id:
                    student_data = sdata
                    scores = student_data.get('scores', {})

                    if scores:
                        overall_marks = [subj_data.get('overall_mark', 0) for subj_data in scores.values() if isinstance(subj_data, dict)]
                        average_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
                        total_marks = sum(overall_marks)
                    else:
                        average_marks = 0
                        total_marks = 0

                    if average_marks >= 70:
                        grade = 'A'
                    elif average_marks >= 60:
                        grade = 'B'
                    elif average_marks >= 50:
                        grade = 'C'
                    elif average_marks >= 40:
                        grade = 'D'
                    else:
                        grade = 'F'

                    status = 'Pass' if average_marks >= 60 else 'Fail'

                    # Restructure subjects as dictionary with score data
                    subjects_with_scores = {}
                    for subject in student_data.get('subjects', []):
                        subj_scores = scores.get(subject, {})
                        if isinstance(subj_scores, dict):
                            subjects_with_scores[subject] = {
                                'tests': subj_scores.get('total_test', 0),
                                'exam': subj_scores.get('total_exam', 0),
                                'total': subj_scores.get('overall_mark', 0)
                            }
                        else:
                            subjects_with_scores[subject] = {'tests': 0, 'exam': 0, 'total': 0}

                    # Build positions across this user's students
                    students_list = []
                    for sid2, sdata2 in user_students.items():
                        sscores2 = sdata2.get('scores', {})
                        if sscores2:
                            overall_marks2 = [subj_data.get('overall_mark', 0) for subj_data in sscores2.values() if isinstance(subj_data, dict)]
                            avg_marks2 = sum(overall_marks2) / len(overall_marks2) if overall_marks2 else 0
                        else:
                            avg_marks2 = 0
                        students_list.append({
                            'first_name': sdata2.get('firstname', ''),
                            'student_id': sid2,
                            'class_name': sdata2.get('classname', ''),
                            'subjects': sscores2,
                            'average_marks': avg_marks2
                        })
                    positions = calculate_positions(students_list)

                    student = {
                        'first_name': student_data.get('firstname', ''),
                        'student_id': sid,
                        'class_name': student_data.get('classname', ''),
                        'number_of_subject': student_data.get('number_of_subject', 0),
                        'subjects': subjects_with_scores,
                        'total_marks': total_marks,
                        'average_marks': average_marks,
                        'Grade': grade,
                        'Status': status
                    }
                    break

    return render_template('search_student.html', student=student, positions=positions)

@app.route('/rank_students')
def rank_students():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    user_students = load_students(user_id)

    class_filter = request.args.get('class', '').strip()

    # Transform data and calculate averages
    students_list = []
    for student_id, student_data in user_students.items():
        scores = student_data.get('scores', {})
        if scores:
            overall_marks = []
            for subj_data in scores.values():
                if isinstance(subj_data, dict):
                    # New format: detailed score object
                    overall_marks.append(subj_data.get('overall_mark', 0))
                elif isinstance(subj_data, (int, float)):
                    # Old format: simple number
                    overall_marks.append(float(subj_data))
            average_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
        else:
            average_marks = 0

        # Determine grade
        if average_marks >= 70:
            grade = 'A'
        elif average_marks >= 60:
            grade = 'B'
        elif average_marks >= 50:
            grade = 'C'
        elif average_marks >= 40:
            grade = 'D'
        else:
            grade = 'F'

        # Determine status
        status = 'Pass' if average_marks >= 60 else 'Fail'

        students_list.append({
            'first_name': student_data.get('firstname', ''),
            'student_id': student_id,
            'class_name': student_data.get('classname', ''),
            'subjects': {subj: subj_data.get('overall_mark', 0) for subj, subj_data in scores.items() if isinstance(subj_data, dict)},
            'average_marks': average_marks,
            'Grade': grade,
            'Status': status
        })

    # Filter by class if specified
    if class_filter:
        students_list = [s for s in students_list if s['class_name'].lower() == class_filter.lower()]

    # Group by class and sort within each class
    ranked = {}
    for student in students_list:
        class_name = student['class_name']
        if class_name not in ranked:
            ranked[class_name] = []
        ranked[class_name].append(student)

    for class_name in ranked:
        ranked[class_name].sort(key=lambda x: x['average_marks'], reverse=True)

    return render_template('rank_students.html', ranked=ranked, class_filter=class_filter)

@app.route('/enter_scores', methods=['GET', 'POST'])
def enter_scores():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    student_id = request.args.get('student_id')
    user_id = session['user_id']

    user_students = load_students(user_id)
    if not student_id or student_id not in user_students:
        flash('Student not found.')
        return redirect(url_for('menu'))

    student = user_students[student_id]

    if request.method == 'POST':
        scores = {}
        invalid_scores = []
        for subject in student['subjects']:
            subj_key = subject.replace(' ', '_')
            subject_scores = {}
            try:
                first_test = float(request.form.get(f'first_test_{subj_key}', 0))
                second_test = float(request.form.get(f'second_test_{subj_key}', 0))
                third_test = float(request.form.get(f'third_test_{subj_key}', 0))
                objective = float(request.form.get(f'objective_{subj_key}', 0))
                theory = float(request.form.get(f'theory_{subj_key}', 0))

                # Validate ranges
                if not (0 <= first_test <= 10): invalid_scores.append(f'{subject}: First test must be 0-10')
                if not (0 <= second_test <= 10): invalid_scores.append(f'{subject}: Second test must be 0-10')
                if not (0 <= third_test <= 10): invalid_scores.append(f'{subject}: Third test must be 0-10')
                if not (0 <= objective <= 30): invalid_scores.append(f'{subject}: Objective must be 0-30')
                if not (0 <= theory <= 40): invalid_scores.append(f'{subject}: Theory must be 0-40')

                # Calculate totals and grades
                calculated_total_test = first_test + second_test + third_test
                calculated_total_exam = objective + theory
                calculated_overall_mark = calculated_total_test + calculated_total_exam
                if not (0 <= calculated_total_test <= 30): invalid_scores.append(f'{subject}: Total test must be 0-30')
                if not (0 <= calculated_total_exam <= 70): invalid_scores.append(f'{subject}: Total exam must be 0-70')
                if not (0 <= calculated_overall_mark <= 100): invalid_scores.append(f'{subject}: Overall mark must be 0-100')

                # Determine test grade based on total test
                if calculated_total_test >= 27:
                    calculated_test_grade = 'A'
                elif calculated_total_test >= 24:
                    calculated_test_grade = 'B'
                elif calculated_total_test >= 20:
                    calculated_test_grade = 'C'
                elif calculated_total_test >= 15:
                    calculated_test_grade = 'D'
                else:
                    calculated_test_grade = 'F'

                # Determine overall grade based on overall mark
                if calculated_overall_mark >= 70:
                    calculated_grade = 'A'
                elif calculated_overall_mark >= 60:
                    calculated_grade = 'B'
                elif calculated_overall_mark >= 50:
                    calculated_grade = 'C'
                elif calculated_overall_mark >= 40:
                    calculated_grade = 'D'
                else:
                    calculated_grade = 'F'

                subject_scores = {
                    'first_test': first_test,
                    'second_test': second_test,
                    'third_test': third_test,
                    'total_test': calculated_total_test,
                    'test_grade': calculated_test_grade,
                    'objective': objective,
                    'theory': theory,
                    'total_exam': calculated_total_exam,
                    'overall_mark': calculated_overall_mark,
                    'grade': calculated_grade
                }
                scores[subject] = subject_scores
            except ValueError:
                invalid_scores.append(f'{subject}: Invalid number format')

        if invalid_scores:
            for error in invalid_scores:
                flash(error, 'error')
            return redirect(url_for('enter_scores', student_id=student_id))

        student['scores'] = scores
        save_student(user_id, student_id, student)
        flash('Scores entered successfully!')
        return redirect(url_for('menu'))

    return render_template('enter_scores.html', student=student, student_id=student_id)

@app.route('/student_report')
def student_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    student_id = request.args.get('student_id')
    user_id = session['user_id']

    user_students = load_students(user_id)
    if not student_id or student_id not in user_students:
        flash('Student not found.')
        return redirect(url_for('view_students'))

    student_data = user_students[student_id]
    scores = student_data.get('scores', {})

    if scores:
        overall_marks = [subj_data.get('overall_mark', 0) for subj_data in scores.values() if isinstance(subj_data, dict)]
        average_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
    else:
        average_marks = 0

    if average_marks >= 70:
        grade = 'A'
    elif average_marks >= 60:
        grade = 'B'
    elif average_marks >= 50:
        grade = 'C'
    elif average_marks >= 40:
        grade = 'D'
    else:
        grade = 'F'

    status = 'Pass' if average_marks >= 60 else 'Fail'

    # Restructure subjects as dictionary with score data
    subjects_with_scores = {}
    for subject in student_data.get('subjects', []):
        subjects_with_scores[subject] = scores.get(subject, {})

    # Build positions across this user's students
    students_list = []
    for sid, sdata in user_students.items():
        sscores = sdata.get('scores', {})
        if sscores:
            overall_marks = [subj_data.get('overall_mark', 0) for subj_data in sscores.values() if isinstance(subj_data, dict)]
            avg_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
        else:
            avg_marks = 0
        students_list.append({
            'first_name': sdata.get('firstname', ''),
            'student_id': sid,
            'class_name': sdata.get('classname', ''),
            'subjects': sscores,
            'average_marks': avg_marks
        })
    positions = calculate_positions(students_list)

    student = {
        'first_name': student_data.get('firstname', ''),
        'student_id': student_id,
        'class_name': student_data.get('classname', ''),
        'number_of_subject': student_data.get('number_of_subject', 0),
        'subjects': subjects_with_scores,  # Now a dict: subject -> score_data
        'average_marks': average_marks,
        'Grade': grade,
        'Status': status
    }

    return render_template('student_report.html', student=student, position=positions.get(student_id), 
                           chart_labels=list(student['subjects'].keys()), 
                           chart_data=[subj_data.get('overall_mark', 0) for subj_data in student['subjects'].values()])

@app.route('/all_students_report')
def all_students_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    user_students = load_students(user_id)
    class_filter = request.args.get('class', '').strip()
    student_id_filter = request.args.get('student_id', '').strip()

    # Transform data to match template expectations
    students_list = []
    for student_id, student_data in user_students.items():
        # Calculate average marks from overall_mark
        scores = student_data.get('scores', {})
        if scores:
            overall_marks = []
            for subj_data in scores.values():
                if isinstance(subj_data, dict):
                    # New format: detailed score object
                    overall_marks.append(subj_data.get('overall_mark', 0))
                elif isinstance(subj_data, (int, float)):
                    # Old format: simple number
                    overall_marks.append(float(subj_data))
            average_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
        else:
            average_marks = 0

        # Determine grade
        if average_marks >= 70:
            grade = 'A'
        elif average_marks >= 60:
            grade = 'B'
        elif average_marks >= 50:
            grade = 'C'
        elif average_marks >= 40:
            grade = 'D'
        else:
            grade = 'F'

        # Determine status
        status = 'Pass' if average_marks >= 60 else 'Fail'

        students_list.append({
            'first_name': student_data.get('firstname', ''),
            'student_id': student_id,
            'class_name': student_data.get('classname', ''),
            'subjects': {
                subj: {
                    'overall_mark': subj_data.get('overall_mark', 0) if isinstance(subj_data, dict) else 0,
                    'grade': subj_data.get('grade', grade) if isinstance(subj_data, dict) else grade,
                    'status': status,
                    'tests': [subj_data.get('first_test', 0), subj_data.get('second_test', 0), subj_data.get('third_test', 0)] if isinstance(subj_data, dict) else [],
                    'cbt': subj_data.get('objective', 0) if isinstance(subj_data, dict) else 0,
                    'theory_exam': subj_data.get('theory', 0) if isinstance(subj_data, dict) else 0,
                    'total_exam': subj_data.get('total_exam', 0) if isinstance(subj_data, dict) else 0,
                    'total_test': subj_data.get('total_test', 0) if isinstance(subj_data, dict) else 0
                }
                for subj, subj_data in scores.items()
            },
            'average_marks': average_marks,
            'Grade': grade,
            'Status': status
        })

    # Apply filters (case-insensitive)
    if class_filter:
        students_list = [
            s for s in students_list
            if s.get('class_name', '').lower() == class_filter.lower()
        ]
    if student_id_filter:
        students_list = [
            s for s in students_list
            if s.get('student_id', '').lower() == student_id_filter.lower()
        ]

    positions = calculate_positions(students_list)

    return render_template(
        'all_students_report.html',
        students=students_list,
        positions=positions,
        now=datetime.now().strftime("%Y-%m-%d %H:%M"),
        class_filter=class_filter,
        student_id_filter=student_id_filter
    )

@app.route('/api/students')
def api_students():
    if 'user_id' not in session:
        return {'error': 'Unauthorized'}, 401

    user_id = session['user_id']
    user_students = load_students(user_id)
    return {'students': user_students}

@app.route('/reports/export_csv')
def reports_export_csv():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    user_students = load_students(user_id)

    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(['Student ID', 'First Name', 'Class', 'Average Marks', 'Grade', 'Status', 'Subjects'])

    # Write data
    for student_id, student_data in user_students.items():
        scores = student_data.get('scores', {})
        if scores:
            overall_marks = [subj_data.get('overall_mark', 0) for subj_data in scores.values() if isinstance(subj_data, dict)]
            average_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
        else:
            average_marks = 0

        if average_marks >= 70:
            grade = 'A'
        elif average_marks >= 60:
            grade = 'B'
        elif average_marks >= 50:
            grade = 'C'
        elif average_marks >= 40:
            grade = 'D'
        else:
            grade = 'F'

        status = 'Pass' if average_marks >= 60 else 'Fail'

        subjects_str = ', '.join(student_data.get('subjects', []))

        writer.writerow([
            student_id,
            student_data.get('firstname', ''),
            student_data.get('classname', ''),
            f"{average_marks:.2f}",
            grade,
            status,
            subjects_str
        ])

    output.seek(0)
    return send_file(BytesIO(output.getvalue().encode('utf-8')), 
                     mimetype='text/csv', 
                     as_attachment=True, 
                     download_name='students_report.csv')

@app.route('/rank/export_csv')
def export_rank_csv():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    # Placeholder: re-use existing CSV export until rank-specific export is implemented
    return redirect(url_for('reports_export_csv', **request.args))

@app.route('/reports/export_xlsx')
def reports_export_xlsx():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        from openpyxl import Workbook
    except ImportError:
        flash('XLSX export requires openpyxl. Please install it.', 'error')
        return redirect(url_for('all_students_report'))

    user_id = session['user_id']
    user_students = load_students(user_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Students Report"

    # Write header
    headers = ['Student ID', 'First Name', 'Class', 'Average Marks', 'Grade', 'Status', 'Subjects']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, (student_id, student_data) in enumerate(user_students.items(), 2):
        scores = student_data.get('scores', {})
        if scores:
            overall_marks = [subj_data.get('overall_mark', 0) for subj_data in scores.values() if isinstance(subj_data, dict)]
            average_marks = sum(overall_marks) / len(overall_marks) if overall_marks else 0
        else:
            average_marks = 0

        if average_marks >= 70:
            grade = 'A'
        elif average_marks >= 60:
            grade = 'B'
        elif average_marks >= 50:
            grade = 'C'
        elif average_marks >= 40:
            grade = 'D'
        else:
            grade = 'F'

        status = 'Pass' if average_marks >= 60 else 'Fail'

        subjects_str = ', '.join(student_data.get('subjects', []))

        ws.cell(row=row_num, column=1, value=student_id)
        ws.cell(row=row_num, column=2, value=student_data.get('firstname', ''))
        ws.cell(row=row_num, column=3, value=student_data.get('classname', ''))
        ws.cell(row=row_num, column=4, value=round(average_marks, 2))
        ws.cell(row=row_num, column=5, value=grade)
        ws.cell(row=row_num, column=6, value=status)
        ws.cell(row=row_num, column=7, value=subjects_str)

    # Save to BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, 
                     download_name='students_report.xlsx')

@app.route('/report_issue', methods=['GET', 'POST'])
def report_issue():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        issue_description = request.form.get('issue_description', '').strip()
        if issue_description:
            report = {
                'user_id': session['user_id'],
                'description': issue_description,
                'timestamp': datetime.now().isoformat(),
                'status': 'unread'
            }
            save_report(report)
            flash('Issue reported successfully! Thank you for your feedback.', 'success')
            return redirect(url_for('menu'))
        else:
            flash('Please describe the issue.', 'error')

    return render_template('report_issue.html')

@app.route('/backup')
def backup():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    return send_file('student_score.db', as_attachment=True, download_name='backup.db')

@app.route('/restore', methods=['GET', 'POST'])
def restore():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        file = request.files.get('file')
        if file and file.filename.endswith('.db'):
            file.save('student_score.db')
            flash('Database restored successfully!', 'success')
            logging.info(f"User {session['user_id']} restored database")
        else:
            flash('Please upload a valid .db file.', 'error')
        return redirect(url_for('menu'))

    return render_template('restore.html')

@app.route('/help')
def help():
    return render_template('help.html')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = True  # Enable debug mode
    app.run(host='0.0.0.0', port=port, debug=debug)


